"""SRD 评估结果 → xlsx。

官网结果页的「导出 Excel」按钮走这里：把一次评估的完整层级（评估 → 领域 → 分组 → 条目）
写成三张工作表，用户拿去归档、传给合作者、或者在 Excel 里自己排序筛选。

**数据源是业务库，不是引擎的 `result.json`** —— 那份文件在任务到终态后连同整个
session 目录一起删掉了（用户上传的综述原文不留），而这里要支持「三个月后从历史列表
里点开再导一次」。

与 `tools/srd-engine/export_excel.py` 不是一回事，别合并：那个是给方法学专家核对
金标准集用的（一次导十对、带评分锚点、带评分矩阵），跑在引擎侧、读 `pair-*.json`；
这个是给终端用户导他自己那一次评估，跑在后端进程里、读 DB。

三张表按「从结论往证据走」排：概览 → 领域 → 条目明细。

中英文各出一版（`lang`）：官网是中英对等的站点，导出件只有中文表头等于英文用户拿到
一份看不懂的存档。表头与档位文案在下面的 `_L` 里成对写死 —— 前台那份在
`i18n/locales/*.json`，两处是同一批词，改判定档位时一起改。
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from module_action.entity.vo.action_vo import SrdAssessmentModel, SrdDomainModel, SrdItemModel

Lang = Literal['zh', 'en']

#: 重复程度档位。与 `srd_engine.aggregate.LEVEL_LABEL_ZH/EN`、前台的 `srd.lvl*` 同一批词。
#: **`None` 单独一档**：领域里一条都没评出来时引擎不查表，兜进任何一档都是在编结论。
_LEVEL: dict[Lang, dict[Any, str]] = {
    'zh': {'none': '无重复', 'low': '低重复', 'mod': '中度重复', 'high': '高度重复', None: '无法判定'},
    'en': {
        'none': 'No duplication',
        'low': 'Low duplication',
        'mod': 'Moderate duplication',
        'high': 'High duplication',
        None: 'Not assessable',
    },
}

#: 条目评分档位。**分越高越重复**（引擎 0.8.0 起），与上面那套百分比色阶同向。
_RATING: dict[Lang, dict[Any, str]] = {
    'zh': {'3': '完全相同', '2': '部分相同', '1': '部分不同', '0': '完全不同', 'unclear': '证据不足', None: '证据不足'},
    'en': {
        '3': 'Identical',
        '2': 'Partly identical',
        '1': 'Partly different',
        '0': 'Completely different',
        'unclear': 'Insufficient evidence',
        None: 'Insufficient evidence',
    },
}

#: 3 分最重复 → 朱砂，0 分最不同 → 青。色阶跟着**重复程度**走而不是跟着分数大小走，
#: 与页面的条目色阶同向，导出的表看着才不别扭。
_RATING_FILL = {
    '3': PatternFill('solid', fgColor='F8CBCB'),
    '2': PatternFill('solid', fgColor='FCE4D6'),
    '1': PatternFill('solid', fgColor='E3EDF6'),
    '0': PatternFill('solid', fgColor='D8EBD5'),
    'unclear': PatternFill('solid', fgColor='F2F2F2'),
}

_HEAD_FILL = PatternFill('solid', fgColor='1F4B44')
_HEAD_FONT = Font(color='FFFFFF', bold=True)
_KEY_FONT = Font(bold=True)
_THIN = Side(style='thin', color='D9D9D9')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_L: dict[Lang, dict[str, str]] = {
    'zh': {
        'sheet_overview': '概览',
        'sheet_domain': '领域',
        'sheet_item': '条目明细',
        'title': 'SRD 系统综述重复性评估',
        'field': '项目',
        'value': '内容',
        'reviewA': '综述 A',
        'reviewB': '综述 B',
        'overall': '整体判定',
        'overallPct': '整体重复度',
        'overallScore': '总得分',
        'scoreFull': '名义满分',
        'reason': '判定理由',
        'provisional': '初步判定',
        'unclearCount': '证据不足条目数',
        'reviewCount': '待人工复核条目数',
        'submitTime': '提交时间',
        'finishTime': '完成时间',
        'provenance': '生成信息',
        'sample': '示例数据（演示用，非真实评估）',
        'yes': '是',
        'no': '否',
        'seq': '序号',
        'domain': '领域',
        'isKey': '关键领域',
        'level': '判定',
        'pct': '重复度',
        'score': '得分',
        'scoreMax': '可评分满分',
        'unclearItems': '证据不足条目',
        'evidence': '证据充分',
        'boundary': '临近分级边界',
        'group': '分组',
        'code': '条目',
        'question': '条目问题',
        'rating': '评分档位',
        'points': '分数',
        'confidence': '把握度',
        'needsReview': '待复核',
        'reviewNote': '复核说明',
        'basis': '判定依据',
        'card': '客观事实',
        'citeA': '综述 A 原文引用',
        'citeATr': '综述 A 中文译文',
        'citeB': '综述 B 原文引用',
        'citeBTr': '综述 B 中文译文',
        'disclaimer': '本工具为辅助判断，结论须由方法学专家确认后方可引用。',
        'scoreNote': '评分 0–3 分，分越高越重复；证据不足的条目不计分，也不进分母。',
    },
    'en': {
        'sheet_overview': 'Overview',
        'sheet_domain': 'Domains',
        'sheet_item': 'Items',
        'title': 'SRD — systematic review duplication assessment',
        'field': 'Field',
        'value': 'Value',
        'reviewA': 'Review A',
        'reviewB': 'Review B',
        'overall': 'Overall verdict',
        'overallPct': 'Overall duplication',
        'overallScore': 'Total score',
        'scoreFull': 'Nominal maximum',
        'reason': 'Rationale',
        'provisional': 'Provisional',
        'unclearCount': 'Items with insufficient evidence',
        'reviewCount': 'Items flagged for human review',
        'submitTime': 'Submitted',
        'finishTime': 'Finished',
        'provenance': 'Run details',
        'sample': 'Sample data (demonstration only, not a real assessment)',
        'yes': 'Yes',
        'no': 'No',
        'seq': 'No.',
        'domain': 'Domain',
        'isKey': 'Key domain',
        'level': 'Verdict',
        'pct': 'Duplication',
        'score': 'Score',
        'scoreMax': 'Scorable maximum',
        'unclearItems': 'Unclear items',
        'evidence': 'Evidence sufficient',
        'boundary': 'Near band boundary',
        'group': 'Group',
        'code': 'Item',
        'question': 'Question',
        'rating': 'Rating',
        'points': 'Points',
        'confidence': 'Confidence',
        'needsReview': 'Needs review',
        'reviewNote': 'Review note',
        'basis': 'Rationale',
        'card': 'Computed facts',
        'citeA': 'Review A — verbatim quote',
        'citeATr': 'Review A — Chinese translation',
        'citeB': 'Review B — verbatim quote',
        'citeBTr': 'Review B — Chinese translation',
        'disclaimer': 'This tool is decision support only; a methodologist must confirm the verdict before it is cited.',
        'scoreNote': 'Items score 0–3; higher means more duplication. Unclear items are excluded from both '
        'the score and the denominator.',
    },
}


def _pick(obj: object, field: str, lang: Lang) -> str:
    """
    取双语字段。

    英文缺失时回落中文，与前台 `useBilingual().pick` 同一条规则 —— 数据侧缺英文时
    不回落会在英文导出件里留一整列空白。

    :param obj: 带 `xxx_zh` / `xxx_en` 两个属性的对象
    :param field: 字段名（不带语言后缀）
    :param lang: 目标语言
    :return: 文本，取不到时空串
    """
    zh = getattr(obj, f'{field}_zh', None) or ''
    en = getattr(obj, f'{field}_en', None) or ''

    return (en or zh) if lang == 'en' else zh


def _cite(raw: str | None, translation: str | None) -> tuple[str, str]:
    """
    拆一条引用成「原文 + 译文」。

    库里 `cite_*_en` 装的是**原语言**的逐字片段（不一定是英文），`cite_*_zh` 是它的
    中文翻译；中文文献两列同值。同值时译文列留空 —— 导出件里并排放两列一模一样的话，
    看的人会以为哪里出了错。

    :param raw: `cite_*_en`，逐字原文
    :param translation: `cite_*_zh`，中文翻译
    :return: (原文, 译文)
    """
    text = (raw or '').strip() or (translation or '').strip()
    tr = (translation or '').strip()

    return text, ('' if tr == text else tr)


def _finish(ws: Any, widths: list[int], *, freeze: str = 'A2', autofilter: bool = True) -> None:
    """
    统一收尾：列宽、表头样式、冻结首行、自动筛选

    :param ws: 工作表
    :param widths: 每列宽度
    :param freeze: 冻结窗格位置
    :param autofilter: 是否给首行加筛选器
    :return: None
    """
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for cell in ws[1]:
        cell.fill, cell.font = _HEAD_FILL, _HEAD_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = freeze
    if autofilter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _sheet_overview(wb: Workbook, a: SrdAssessmentModel, lang: Lang) -> None:
    """
    概览表：一行一个字段，看完就知道结论是什么、可不可信

    :param wb: 工作簿
    :param a: 评估结果
    :param lang: 语言
    :return: None
    """
    lb = _L[lang]
    ws = wb.create_sheet(lb['sheet_overview'])
    ws.append([lb['field'], lb['value']])

    score = '' if a.overall_score_max is None else f'{a.overall_score_sum or 0}/{a.overall_score_max}'
    rows: list[tuple[str, Any]] = [
        (lb['reviewA'], _pick(a, 'review_a_title', lang)),
        (lb['reviewB'], _pick(a, 'review_b_title', lang)),
        (lb['overall'], _LEVEL[lang][a.overall_level]),
        # 判定不出来时百分比是 0，照直写「0%」等于替模型下了「两篇毫无重叠」这个结论
        (lb['overallPct'], '' if a.overall_level is None else f'{a.overall_pct or 0}%'),
        (lb['overallScore'], score),
        (lb['scoreFull'], a.overall_score_max_full or ''),
        (lb['unclearCount'], a.unclear_count or 0),
        (lb['reviewCount'], a.review_count or 0),
        (lb['provisional'], lb['yes'] if a.provisional == '1' else lb['no']),
        (lb['reason'], _pick(a, 'overall_reason', lang)),
    ]
    if a.is_sample == '1':
        rows.insert(0, (lb['provenance'], lb['sample']))
    else:
        rows.extend([
            (lb['submitTime'], a.create_time.strftime('%Y-%m-%d %H:%M') if a.create_time else ''),
            (lb['finishTime'], a.finish_time.strftime('%Y-%m-%d %H:%M') if a.finish_time else ''),
            # 导出件是要存档的：哪个模型、哪版引擎跑出来的，是日后复现这份结论的唯一线索。
            # 页面上刻意不显示（用户不关心），存档件里必须有。
            (lb['provenance'], ' · '.join(
                p for p in (
                    a.model_name or '',
                    f'{a.llm_calls} calls' if a.llm_calls else '',
                    f'{round(a.seconds)}s' if a.seconds else '',
                    a.engine_version or '',
                ) if p
            )),
        ])
    rows.extend([('', ''), ('', lb['scoreNote']), ('', lb['disclaimer'])])

    for name, value in rows:
        ws.append([name, value])
    for row in ws.iter_rows(min_row=2, max_col=1):
        row[0].font = _KEY_FONT
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(vertical='top', wrap_text=True)
    _finish(ws, [22, 96], autofilter=False)


def _sheet_domains(wb: Workbook, a: SrdAssessmentModel, lang: Lang) -> None:
    """
    领域表：四行，看每个领域的百分比是怎么算出来的

    :param wb: 工作簿
    :param a: 评估结果
    :param lang: 语言
    :return: None
    """
    lb = _L[lang]
    ws = wb.create_sheet(lb['sheet_domain'])
    ws.append([
        lb['seq'], lb['domain'], lb['isKey'], lb['level'], lb['pct'],
        lb['score'], lb['scoreMax'], lb['scoreFull'], lb['unclearItems'], lb['evidence'], lb['boundary'],
    ])
    for d in a.domains:
        ws.append([
            d.seq,
            _pick(d, 'name', lang),
            lb['yes'] if d.is_key == '1' else lb['no'],
            _LEVEL[lang][d.level],
            '' if d.level is None else f'{d.pct or 0}%',
            d.score_sum if d.score_sum is not None else '',
            d.score_max if d.score_max is not None else '',
            d.score_max_full if d.score_max_full is not None else '',
            d.unclear_count or 0,
            lb['yes'] if d.evidence_sufficient == '1' else lb['no'],
            lb['yes'] if d.near_boundary == '1' else lb['no'],
        ])
    _finish(ws, [6, 26, 10, 14, 10, 8, 12, 12, 12, 10, 14])


def _item_rows(a: SrdAssessmentModel, lang: Lang) -> list[tuple[SrdDomainModel, str, SrdItemModel]]:
    """
    把三层结构摊平成条目行

    :param a: 评估结果
    :param lang: 语言
    :return: [(领域, 分组名, 条目)]
    """
    return [
        (d, f'{g.code} {_pick(g, "name", lang)}'.strip(), it)
        for d in a.domains
        for g in d.groups
        for it in g.items
    ]


def _sheet_items(wb: Workbook, a: SrdAssessmentModel, lang: Lang) -> None:
    """
    条目明细：34 行，人工复核的主表 —— 判定依据与两篇的逐字引用并排放，逐条判对错

    :param wb: 工作簿
    :param a: 评估结果
    :param lang: 语言
    :return: None
    """
    lb = _L[lang]
    ws = wb.create_sheet(lb['sheet_item'])
    ws.append([
        lb['domain'], lb['group'], lb['code'], lb['question'], lb['rating'], lb['points'],
        lb['confidence'], lb['needsReview'], lb['reviewNote'], lb['basis'], lb['card'],
        lb['citeA'], lb['citeATr'], lb['citeB'], lb['citeBTr'],
    ])
    for d, group_name, it in _item_rows(a, lang):
        cite_a, cite_a_tr = _cite(it.cite_a_en, it.cite_a_zh)
        cite_b, cite_b_tr = _cite(it.cite_b_en, it.cite_b_zh)
        ws.append([
            _pick(d, 'name', lang),
            group_name,
            it.code or '',
            _pick(it, 'question', lang),
            _RATING[lang][it.rating],
            it.score if it.score is not None else '',
            it.confidence or '',
            lb['yes'] if it.needs_review == '1' else '',
            it.review_note or '',
            _pick(it, 'basis', lang),
            it.evidence_card or '',
            cite_a, cite_a_tr, cite_b, cite_b_tr,
        ])
        ws.cell(row=ws.max_row, column=5).fill = _RATING_FILL.get(it.rating or 'unclear', _RATING_FILL['unclear'])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = _BORDER
    _finish(ws, [16, 22, 8, 44, 14, 7, 10, 9, 22, 48, 30, 48, 40, 48, 40], freeze='D2')


def build_assessment_xlsx(assessment: SrdAssessmentModel, lang: Lang = 'zh') -> bytes:
    """
    把一次 SRD 评估导成 xlsx 的二进制内容

    :param assessment: 完整评估结果（评估 → 领域 → 分组 → 条目）
    :param lang: 导出语言（zh / en）
    :return: xlsx 文件字节
    """
    lang = lang if lang in _L else 'zh'
    wb = Workbook()
    # Workbook() 自带一张空的 Sheet，不删的话导出件第一张表是空白页
    wb.remove(wb.active)
    _sheet_overview(wb, assessment, lang)
    _sheet_domains(wb, assessment, lang)
    _sheet_items(wb, assessment, lang)
    wb.properties.title = _L[lang]['title']

    buffer = io.BytesIO()
    wb.save(buffer)

    return buffer.getvalue()
