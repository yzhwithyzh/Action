"""把评估结果导成一份便于人工核对的 Excel。

    python export_excel.py                      # 读 out/，写 out/srd-结果.xlsx
    python export_excel.py out-0.6.0            # 指定结果目录

只读 `pair-*.json`，不调模型 —— 想换个排版重导随时可以跑，不花钱。

三个工作表，按「从结论往证据走」的顺序排：

1. **汇总**   —— 一对一行，看整体得分、整体判定与四个领域的档位
2. **领域**   —— 一对四行，看每个领域的得分、满分与重复百分比是怎么来的
3. **条目明细** —— 一对 34 行（共 340 行），**这是人工核对的主表**：
   条目原题 + 评分 + 理由 + 两篇的逐字引用并排放，逐条判对错
4. **评分矩阵** —— 34 行 × 10 列，一眼看全部配对在同一条目上的分数，找系统性偏差用

条目明细里刻意把「评分锚点」也带上：核对的人要判断模型打得对不对，
就得知道这条的 0 分 / 3 分分别该长什么样，否则只能凭感觉。
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from srd_engine.checklist import ALL_ITEMS, DOMAINS, criteria_of
from srd_engine.schemas import RATING_LABEL_ZH, SCORE_PER_ITEM, AssessmentResult

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

LEVEL_ZH = {'none': '无重复', 'low': '低度重复', 'mod': '中度重复', 'high': '高度重复', None: '证据不足'}
# 底色跟着**重复程度**走，不跟着分数大小走：完全相同最红，完全不同最绿。
# 误判「重复」的代价最高，所以红端最扎眼。0.8.0 翻转分数后，红端从 0 分挪到了 3 分。
RATING_FILL = {
    '3': PatternFill('solid', fgColor='F8CBCB'),        # 红：完全相同
    '2': PatternFill('solid', fgColor='FCE4D6'),        # 橙：部分相同
    '1': PatternFill('solid', fgColor='EAF3DE'),        # 浅绿：部分不同
    '0': PatternFill('solid', fgColor='D8EBD5'),        # 绿：完全不同
    'unclear': PatternFill('solid', fgColor='F2F2F2'),  # 灰：证据不足
}
HEAD_FILL = PatternFill('solid', fgColor='305496')
HEAD_FONT = Font(color='FFFFFF', bold=True)
GROUP_OF = {it.code: (d, g) for d in DOMAINS for g in d.groups for it in g.items}


def _style_header(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def sheet_summary(wb: Workbook, results: list[tuple[int, AssessmentResult]]) -> None:
    ws = wb.create_sheet('汇总')
    ws.append(['配对', '综述A', '综述B', '整体判定', '总得分', '名义满分', '重复度%', '待人工复核',
               '领域1 研究主题(关键)', '领域2 研究方法', '领域3 研究结果(关键)', '领域4 研究质量',
               '0分', '1分', '2分', '3分', '证据不足', '结论依据', '调用', '入token', '出token'])
    for no, r in results:
        by_seq = {d.seq: d for d in r.domains}
        cells = [f'{d.score_sum}/{d.score_max}　{LEVEL_ZH[d.level]} {d.pct}%' if d.level else '证据不足'
                 for d in (by_seq[1], by_seq[2], by_seq[3], by_seq[4])]
        ratings = [it.effective_rating for it in r.items]
        ws.append([
            no, r.review_a_title[:60], r.review_b_title[:60],
            LEVEL_ZH[r.overall_level] + ('（初步）' if r.provisional else ''),
            f'{r.overall_score_sum}/{r.overall_score_max}', r.overall_score_max_full, r.overall_pct,
            r.review_count, *cells,
            *(ratings.count(k) for k in ('0', '1', '2', '3', 'unclear')),
            r.overall_reason_zh, r.llm_calls, r.token_in, r.token_out,
        ])
    _style_header(ws, [6, 36, 36, 14, 10, 10, 9, 12, 24, 22, 24, 22, 6, 6, 6, 6, 10, 70, 8, 10, 10])
    for row in ws.iter_rows(min_row=2):
        row[17].alignment = Alignment(wrap_text=True, vertical='top')


def sheet_domains(wb: Workbook, results: list[tuple[int, AssessmentResult]]) -> None:
    ws = wb.create_sheet('领域')
    ws.append(['配对', '领域', '关键领域', '档位', '重复度%', '得分', '可评分满分', '名义满分',
               '证据不足条数', '百分比算法', '证据是否充分', '临界'])
    for no, r in results:
        for d in r.domains:
            ws.append([
                no, d.name_zh, '是' if d.is_key else '', LEVEL_ZH[d.level], d.pct,
                d.score_sum, d.score_max, d.score_max_full, d.unclear_count,
                f'{d.score_sum}/{d.score_max} = {d.pct}%'
                f'（证据不足的 {d.unclear_count} 条连同其 {3 * d.unclear_count} 分一并剔出）'
                if d.score_max else '无可评分条目',
                '是' if d.evidence_sufficient else '否 ← 结论仅供参考',
                '是 ← 一两分之差就跳档' if d.near_boundary else '',
            ])
    _style_header(ws, [6, 20, 10, 12, 10, 8, 12, 10, 12, 52, 22, 22])


def sheet_items(wb: Workbook, results: list[tuple[int, AssessmentResult]]) -> None:
    ws = wb.create_sheet('条目明细')
    ws.append(['配对', '领域', '分组', '条目', '评估问题', '评分', '评分档位', '人工复核(评分)',
               '评分理由', '综述A 原文引用', '综述B 原文引用', '把握度', '待复核',
               '评分锚点（3 分 / 0 分 / 证据不足）', '程序算出的客观事实'])
    for no, r in results:
        for it in r.items:
            d, g = GROUP_OF[it.code]
            c = criteria_of(it.code)
            criteria = (f'3 分（完全相同）：{c.get("dup_when", "").strip()}\n'
                        f'0 分（完全不同）：{c.get("diff_when", "").strip()}\n'
                        f'证据不足：{c.get("unclear_when", "").strip()}\n'
                        f'2 分 / 1 分：两个锚点之间，看差异会不会改变临床或方法学解读')
            ws.append([
                no, d.name_zh, f'{g.code}. {g.name_zh}', it.code, it.question_zh,
                '' if it.score is None else it.score, RATING_LABEL_ZH[it.effective_rating],
                '',                                             # 人工复核列留空供填写
                it.reason_zh, it.cite_a, it.cite_b,
                it.confidence, '是' if it.needs_review else '',
                criteria, it.evidence_card,
            ])
            fill = RATING_FILL[it.effective_rating]
            ws.cell(row=ws.max_row, column=6).fill = fill
            ws.cell(row=ws.max_row, column=7).fill = fill
    _style_header(ws, [6, 16, 22, 8, 40, 6, 10, 12, 55, 45, 45, 8, 8, 60, 40])
    for row in ws.iter_rows(min_row=2):
        for i in (4, 8, 9, 10, 13, 14):
            row[i].alignment = Alignment(wrap_text=True, vertical='top')


def sheet_matrix(wb: Workbook, results: list[tuple[int, AssessmentResult]]) -> None:
    """34 行 × N 列的评分矩阵：横着看一对，竖着看同一条目在所有配对上的分数。

    竖着看是这张表存在的理由 —— 某个条目十对全打 3 分（判为完全相同），多半是口径写松了
    或抽取没喂到料，这种系统性偏差在「条目明细」那种一对一对翻的排版里根本看不出来。
    """
    ws = wb.create_sheet('评分矩阵')
    ws.append(['条目', '评估问题', *[f'配对{no}' for no, _ in results], '均分', '打3分的对数'])
    by_pair = [{it.code: it for it in r.items} for _, r in results]
    for item in ALL_ITEMS:
        cells = [by_pair[i].get(item.code) for i in range(len(results))]
        scores = [c.score for c in cells if c is not None and c.score is not None]
        ws.append([
            item.code, item.question_zh,
            *['' if c is None or c.score is None else c.score for c in cells],
            round(sum(scores) / len(scores), 2) if scores else '',
            sum(1 for s in scores if s == SCORE_PER_ITEM),   # 满分 = 判为「完全相同」
        ])
        for col, c in enumerate(cells, start=3):
            if c is not None:
                ws.cell(row=ws.max_row, column=col).fill = RATING_FILL[c.effective_rating]
    _style_header(ws, [8, 52, *[8] * len(results), 8, 12])
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical='top')


def sheet_readme(wb: Workbook, out_dir: Path, results: list[tuple[int, AssessmentResult]]) -> None:
    ws = wb.create_sheet('说明', 0)
    r0 = results[0][1]
    rows = [
        ['SRD 系统综述重复性评估 · 人工核对表', ''],
        ['', ''],
        ['结果目录', str(out_dir)],
        ['配对数', len(results)],
        ['引擎版本', r0.engine_version],
        ['提示词版本', r0.prompt_version],
        ['判定口径版本', r0.criteria_version],
        ['模型', r0.model],
        ['判定粒度', f'{r0.judge_granularity}（34 条目一次调用判完）'],
        ['', ''],
        ['评分口径', '每个条目 0–3 分，分越高越重复：'],
        ['', '3 = 完全相同　2 = 部分相同　1 = 部分不同　0 = 完全不同　（空 = 证据不足，不计分）'],
        ['', '领域满分 = 3 × 条目数：领域1 /24、领域2 /18、领域3 /42、领域4 /18，合计 /102'],
        ['', '领域重复度% = 得分 ÷ 可评分满分 × 100'],
        ['', '重复程度分档（表 2）：0–25% 无重复｜26–50% 低重复｜51–75% 中度重复｜76–100% 高度重复'],
        ['', '整体判定（表 3）：关键领域（主题、结果）× 非关键领域（方法、质量）查表，不是简单平均'],
        ['', ''],
        ['怎么核对', '打开「条目明细」表，逐行看：'],
        ['', '① 读 E 列「评估问题」和 N 列「评分锚点」，明确这一条 3 分和 0 分各该长什么样'],
        ['', '② 看 J/K 列两篇的原文引用，自己打一次分'],
        ['', '③ 与 F 列「评分」比对，把你的分数填进 H 列「人工复核(评分)」'],
        ['', '④ I 列「评分理由」是模型的说法，可用来定位它错在哪一步'],
        ['', ''],
        ['需要重点看的', '· F 列标红的 3 分（判为完全相同）—— 误判重复的代价最高'],
        ['', '· M 列标「是」的（模型自己标了待复核）'],
        ['', '· 「评分矩阵」表里竖着看某条目十对全是同一个分 —— 多半是口径或抽取的系统性问题'],
        ['', '· 「汇总」表里带「（初步）」的配对 —— 关键领域证据不足，结论本就不可靠'],
        ['', '· 「领域」表里「临界」列标「是」的 —— 一两分之差就会跳档'],
        ['', ''],
        ['⚠ 已知局限', '· 引用真伪不再由程序校验（0.4.0 移除），J/K 列的引用可能是模型编的，'],
        ['', '  核对时请回原文确认'],
        ['', '· 同一配置重跑，条目级判定约有 5% 的自然抖动；四档评分比原来的两档更容易抖 1 分'],
        ['', '· 「证据不足」连同它那 3 分一并剔出分母，所以 unclear 多的配对，分母很小'],
        ['', '· 2 分 / 1 分的中间档目前只有通用口径（差异会不会改变临床或方法学解读），'],
        ['', '  尚未逐条写专属锚点，也未经方法学专家评审 —— 这两档的判定请重点复核'],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 90
    ws['A1'].font = Font(bold=True, size=14)
    for row in ws.iter_rows(min_row=1):
        row[1].alignment = Alignment(wrap_text=True, vertical='top')


def main(argv: list[str]) -> int:
    out_dir = Path(argv[0]) if argv else Path('out')
    files = sorted(out_dir.glob('pair-*.json'))
    if not files:
        print(f'{out_dir} 下没有 pair-*.json', file=sys.stderr)
        return 1

    results = [(int(f.stem.split('-')[1]),
                AssessmentResult.model_validate_json(f.read_text(encoding='utf-8')))
               for f in files]
    results.sort(key=lambda x: x[0])

    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, results)
    sheet_domains(wb, results)
    sheet_items(wb, results)
    sheet_matrix(wb, results)
    sheet_readme(wb, out_dir, results)

    path = out_dir / 'SRD-评估结果-人工核对.xlsx'
    wb.save(path)
    n_items = sum(len(r.items) for _, r in results)
    ratings = [it.effective_rating for _, r in results for it in r.items]
    print(f'已写出 {path}')
    print(f'  {len(results)} 对，条目明细 {n_items} 行')
    print('  评分分布：' + '　'.join(
        f'{k}分 {ratings.count(k)}' if k != 'unclear' else f'证据不足 {ratings.count(k)}'
        for k in ('0', '1', '2', '3', 'unclear')
    ))
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv[1:]))
